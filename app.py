from flask import Flask, request, render_template_string, Response
from openpyxl import load_workbook, Workbook
import os
import io

app = Flask(__name__)

base_path = os.path.dirname(__file__)
arquivos_cds = {
    "HUB": os.path.join(base_path, "cd1.xlsx"),
    "DUTRA": os.path.join(base_path, "CD2.xlsx"),
    "JACANA": os.path.join(base_path, "CD3.xlsx"),
    "LESTE": os.path.join(base_path, "CD4.xlsx"),
    "JABAQUARA": os.path.join(base_path, "CD5.xlsx"),
    "CENTRO": os.path.join(base_path, "CD6.xlsx")
}

def carregar_dados():
    dados = []
    for origem, caminho in arquivos_cds.items():
        if os.path.exists(caminho):
            wb = load_workbook(caminho)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                codigo = row[0]       # Coluna A
                nome = row[1]         # Coluna B
                qtd_disponivel = row[7]  # Coluna H
                if codigo and nome:
                    dados.append({
                        "codigo": str(codigo),
                        "nome": nome,
                        "disponivel": qtd_disponivel,
                        "origem": origem
                    })
    return dados

def aplicar_filtros(dados, codigo, nome, qtd_min, qtd_max, origem):
    if codigo:
        dados = [d for d in dados if d["codigo"] == codigo]
    if nome:
        dados = [d for d in dados if nome.lower() in d["nome"].lower()]
    if qtd_min:
        dados = [d for d in dados if d["disponivel"] is not None and d["disponivel"] >= int(qtd_min)]
    if qtd_max:
        dados = [d for d in dados if d["disponivel"] is not None and d["disponivel"] <= int(qtd_max)]
    if origem:
        dados = [d for d in dados if d["origem"].lower() == origem.lower()]
    return dados

@app.route("/")
def index():
    codigo = request.args.get("codigo")
    nome = request.args.get("nome")
    qtd_min = request.args.get("qtd_min")
    qtd_max = request.args.get("qtd_max")
    origem = request.args.get("origem")

    dados = carregar_dados()
    dados = aplicar_filtros(dados, codigo, nome, qtd_min, qtd_max, origem)

    html = """
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
        <meta charset="UTF-8">
        <title>Consulta Estoques</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                background: linear-gradient(135deg, #1f4037, #99f2c8);
                margin: 0;
                padding: 20px;
                color: #333;
            }
            h1 {
                text-align: center;
                color: #fff;
                margin-bottom: 20px;
            }
            form {
                background: #fff;
                padding: 15px;
                border-radius: 8px;
                margin-bottom: 20px;
                box-shadow: 0 4px 8px rgba(0,0,0,0.1);
                display: flex;
                flex-wrap: wrap;
                gap: 10px;
                justify-content: center;
            }
            form input[type="text"],
            form input[type="number"],
            form select {
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 6px;
            }
            form input[type="submit"],
            form button {
                background: #1f4037;
                color: #fff;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                cursor: pointer;
                transition: 0.3s;
            }
            form input[type="submit"]:hover,
            form button:hover {
                background: #0d2c22;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                background: #fff;
                border-radius: 8px;
                overflow: hidden;
                box-shadow: 0 4px 8px rgba(0,0,0,0.1);
            }
            th {
                background: #1f4037;
                color: #fff;
                padding: 12px;
                text-align: center;
            }
            td {
                padding: 10px;
                text-align: center;
                border-bottom: 1px solid #ddd;
            }
            tr:hover {
                background: #f2f2f2;
            }
        </style>
    </head>
    <body>
        <h1>CONTROLE DE ESTOQUES</h1>
        <form method="get">
            Código: <input type="text" name="codigo" value="{{ request.args.get('codigo','') }}">
            Nome: <input type="text" name="nome" value="{{ request.args.get('nome','') }}">
            Quantidade mínima: <input type="number" name="qtd_min" value="{{ request.args.get('qtd_min','') }}">
            Quantidade máxima: <input type="number" name="qtd_max" value="{{ request.args.get('qtd_max','') }}">
            Origem:
            <select name="origem">
                <option value=""></option>
                <option value="HUB" {% if request.args.get('origem')=='HUB' %}selected{% endif %}>HUB</option>
                <option value="DUTRA" {% if request.args.get('origem')=='DUTRA' %}selected{% endif %}>DUTRA</option>
                <option value="JAÇANA" {% if request.args.get('origem')=='JAÇANA' %}selected{% endif %}>JAÇANA</option>
                <option value="LESTE" {% if request.args.get('origem')=='LESTE' %}selected{% endif %}>LESTE</option>
                <option value="JABAQUARA" {% if request.args.get('origem')=='JABAQUARA' %}selected{% endif %}>JABAQUARA</option>
                <option value="CENTRO" {% if request.args.get('origem')=='CENTRO' %}selected{% endif %}>CENTRO</option>
            </select>
            <input type="submit" value="Filtrar">
            <button type="submit" formaction="/exportar">Exportar Excel</button>
        </form>
        <table>
            <tr><th>Código</th><th>Nome</th><th>Disponível</th><th>Origem</th></tr>
            {% for d in dados %}
            <tr>
                <td>{{ d.codigo }}</td>
                <td>{{ d.nome }}</td>
                <td>{{ d.disponivel }}</td>
                <td>{{ d.origem }}</td>
            </tr>
            {% endfor %}
        </table>
    </body>
    </html>
    """
    return render_template_string(html, dados=dados)

@app.route("/exportar")
def exportar():
    codigo = request.args.get("codigo")
    nome = request.args.get("nome")
    qtd_min = request.args.get("qtd_min")
    qtd_max = request.args.get("qtd_max")
    origem = request.args.get("origem")

    dados = carregar_dados()
    dados = aplicar_filtros(dados, codigo, nome, qtd_min, qtd_max, origem)

    # Criar workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoques Filtrados"

    # Cabeçalho
    ws.append(["Código", "Nome", "Disponível", "Origem"])

    # Dados
    for d in dados:
        ws.append([d["codigo"], d["nome"], d["disponivel"], d["origem"]])

    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(output,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment;filename=estoques_filtrados.xlsx"})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
